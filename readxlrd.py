import openpyxl

def get_data():
    data = [] #申明list
    wb = openpyxl.load_workbook('test.xlsx') #导入工作表
    sheetnames = wb.sheetnames #获得表单名字
    sheet = wb[sheetnames[0]] #从工作表中提取某一表单

    # 修改
    # sheet['B3'] = '123'
    # sheet['J2'] = '=AVERAGE(A2:I2)'

    # 获得最大列和最大行
    print(sheet.max_row)
    print(sheet.max_column)

    # for rowNum in range(1,24):
    #     data.append(sheet.cell(row=rowNum,column=1).value) #获得数据
    # print(data) #读出数据
    # 因为按行，所以返回A1, B1, C1这样的顺序
    for row in sheet.rows:
        temp = []
        for cell in row:
            temp.append(cell.value)
        data.append(temp)

    # A1, A2, A3这样的顺序
    # for column in sheet.columns:
    #     for cell in column:
    #         print(cell.value)

    # 保存
    # wb.save('test.xlsx')
    return data

if __name__ == '__main__':
    data = get_data()
    print(data)